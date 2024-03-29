"""
.. module:: generate_machine_cim.py
   :license: GPL/CeCIL
   :platform: Unix, Windows
   :synopsis: Converts CMIP6 machine spreadsheets to a machine CIM document.

.. moduleauthor::
   Sadie Bartholomew <sadie.bartholomew@ncas.ac.uk>

"""

import json
import os
from pprint import pprint

from openpyxl import load_workbook

import pyesdoc
from pyesdoc.ontologies.cim import v2 as cim

from lib.utils import logger, vocabs


# Define command line argument parser.
_ARGS = argparse.ArgumentParser(
    "Generates a CMIP6 CIM v2.2 document for every machine of the institute.")
_ARGS.add_argument(
    "--spreadsheet",
    help="Path to the institute's CMIP6 machine worksheet.",
    dest="spreadsheet_filepath",
    type=str
)
_ARGS.add_argument(
    "--io-dir",
    help="Path to a directory into which documents will be written.",
    dest="io_dir",
    type=str
)
_ARGS.add_argument(
    "--institution-id",
    help="An institution identifier",
    dest="institution_id",
    type=str
)
_ARGS = _ARGS.parse_args()


# Validate command line options.
if not os.path.isfile(_ARGS.spreadsheet_filepath):
    raise ValueError("Spreadsheet file does not exist")
if not os.path.isdir(_ARGS.io_dir):
    raise ValueError(
        "Archive directory does not exist: {}".format(_ARGS.io_dir))


INSTITUTE = _ARGS.institution_id
WS_IN_PATH = _ARGS.spreadsheet_filepath
CIM_OUT_PATH = _ARGS.io_dir

LABEL_COLUMN = 0  # i.e. index in A-Z of columns as tuple, so "A"
INPUT_COLUMN = 1  # i.e. "B"

EMPTY_CELL_MARKER = "NO CELL VALUE SPECIFIED"

# Usually <500, but in case of much cell copying and all exps and models listed
MAX_ROW = 3100
MAX_NUMBER_MODELS_PER_INSTITUTE = 141  # strict upper limit + 1
MAX_NUMBER_MIPS = 25  # ditto to above

# Tuple keys give digits corresopnding to question labels to match to user
# inputs, e.g. (1, 1, 1) -> question "1.1.1" or "1.1.1 *", values are offsets,
# where a "SPECIAL CASE:" string value indicates extra rows may have been
# created by the user, and is to be dealt with on a case-by-case basis:
WS_QUESTIONS_TO_INPUT_CELLS_MAPPING = {
    # Identity:
    (1, 1, 1): 2,
    (1, 1, 2): 5,
    # General properties:
    (1, 2, 1): 4,
    (1, 2, 2): 2,
    (1, 2, 3): "SPECIAL CASE: 4+",
    (1, 2, 4): "SPECIAL CASE: 4+6",
    (1, 2, 5): 2,
    # Vendor information:
    (1, 3, 1): 4,
    (1, 3, 2): 2,

    # Compute pools (second pool to be added by mirroring the first, later)
    # Compute pool 1:
    (1, 4, 1, 1): 2,
    (1, 4, 1, 2): 2,
    (1, 4, 1, 3): 4,
    (1, 4, 1, 4): 2,
    (1, 4, 1, 5): 2,
    (1, 4, 1, 6): 2,
    (1, 4, 1, 7): 2,
    (1, 4, 1, 8, 1, 1): 2,
    (1, 4, 1, 8, 1, 2): 2,
    (1, 4, 1, 8, 1, 3): 4,
    (1, 4, 1, 8, 2, 1): 2,
    (1, 4, 1, 8, 2, 2): 2,
    (1, 4, 1, 8, 2, 3): 4,
    (1, 4, 1, 8, 3, 1): 2,
    (1, 4, 1, 8, 3, 2): 2,
    (1, 4, 1, 8, 3, 3): 4,
    (1, 4, 1, 9): 2,
    (1, 4, 1, 10): 2,
    (1, 4, 1, 11): 2,
    (1, 4, 1, 12): 2,
    (1, 4, 1, 13): 2,

    # Storage pools (second pool to be added by mirroring the first, later)
    # Storage pool 1:
    (1, 5, 1, 1): 2,
    (1, 5, 1, 2): "SPECIAL CASE: 4+",
    (1, 5, 1, 3): 2,
    (1, 5, 1, 4): 4,
    (1, 5, 1, 5): 4,

    # Interconnect:
    (1, 6, 1): 2,
    (1, 6, 2): 2,
    (1, 6, 3): 2,
    (1, 6, 4): 4,
    # Benchmark performance:
    (1, 7, 1): 2,
    (1, 7, 2): 2,

    # Applicability...
    # Applicable models:
    (1, 8, 1): 2,
    # All (1, 8, 2, N) are processed in below
    # Applicable experiments:
    (1, 9, 1): 2,
    # All (1, 9, 2, N) are processed in below
}

# Denotes questions which map to CIM components of non-string type which
# must be converted from the string inputs extracted from the spreadsheet.
# TODO manage units from CIM.
WS_QUESTIONS_WITH_NON_STRING_TYPE = {
    (1, 4, 1, 5): int,
    (1, 4, 1, 6): int,  # for CIM v2.0 only, float at 2.2
    (1, 4, 1, 7): int,
    (1, 4, 1, 8, 1, 2): float,
    (1, 4, 1, 8, 2, 2): float,
    (1, 4, 1, 8, 3, 2): float,
    (1, 4, 1, 9): int,
    (1, 4, 1, 12): float,
    (1, 4, 1, 13): int,
    (1, 4, 2, 5): int,
    (1, 4, 2, 6): int,  # for CIM v2.0 only, float at 2.2
    (1, 4, 2, 7): int,
    (1, 4, 2, 8, 1, 2): float,
    (1, 4, 2, 8, 2, 2): float,
    (1, 4, 2, 8, 3, 2): float,
    (1, 4, 2, 9): int,
    (1, 4, 2, 12): float,
    (1, 4, 2, 13): int,
    (1, 5, 1, 2): float,
    (1, 5, 2, 2): float,
    (1, 7, 1): float,
    (1, 7, 2): float,
}

COMPUTE_POOL_2_Q_NOS = (1, 4, 2)
STORAGE_POOL_2_Q_NOS = (1, 5, 2)

# TODO this is for CIM V2.0 only, TODO CIM V2.0 -> V2.2
WS_QUESTIONS_WITH_ASSOCIATIONS = {
    (1, 2, 1): "Party",
    #(1, 2, 3): "OnlineResource",
    (1, 2, 4): "TimePeriod",
    (1, 3, 1): "Party",
    (1, 4, 1, 6): "StorageVolume",
    (1, 4, 2, 6): "StorageVolume",
    # (1, 5, 1, 4): "StorageSystems",  # validates to a SS string by nature of
    # (1, 5, 2, 4): "StorageSystems",  # ... the spreadsheet enum, so not req.
    (1, 5, 1, 5): "Party",
    (1, 5, 2, 5): "Party",
}

QUESTIONS_TO_CIM_MAPPING = {
    # Identity:
    (1, 1, 1): ("name",),
    (1, 1, 2): ("partition", "name"),
    # General properties:
    (1, 2, 1): ("institution",),
    (1, 2, 2): ("description",),
    (1, 2, 3): ("online_documentation",),
    (1, 2, 4): ("when_used",),
    ### (1, 2, 5): ("operating_system",),  # TODO CIM V2.0 -> V2.2
    # Vendor information:
    (1, 3, 1): ("vendor",),
    (1, 3, 2): ("model_number",),
    # Compute pools...
    # Compute pool 1:
    (1, 4, 1, 1): ("compute_pools", "name"),
    (1, 4, 1, 2): ("compute_pools", "description"),
    ### (1, 4, 1, 3): ("compute_pools", "vendor"),  # TODO CIM V2.0 -> V2.2
    (1, 4, 1, 4): ("compute_pools", "model_number"),
    (1, 4, 1, 5): ("compute_pools", "number_of_nodes"),
    (1, 4, 1, 6): ("compute_pools", "memory_per_node"),
    (1, 4, 1, 7): ("compute_pools", "compute_cores_per_node"),
    # TODO CIM V2.0 -> V2.2, NIC component:
    ### (1, 4, 1, 8, 1, 1): ("compute_pools", "nic", "name"),
    ### (1, 4, 1, 8, 1, 2): ("compute_pools", "nic", "bandwidth"),
    ### (1, 4, 1, 8, 1, 3): ("compute_pools", "nic", "vendor"),
    ### (1, 4, 1, 8, 2, 1): ("compute_pools", "nic", "name"),
    ### (1, 4, 1, 8, 2, 2): ("compute_pools", "nic", "bandwidth"),
    ### (1, 4, 1, 8, 2, 3): ("compute_pools", "nic", "vendor"),
    ### (1, 4, 1, 8, 3, 1): ("compute_pools", "nic", "name"),
    ### (1, 4, 1, 8, 3, 2): ("compute_pools", "nic", "bandwidth"),
    ### (1, 4, 1, 8, 3, 3): ("compute_pools", "nic", "vendor"),
    (1, 4, 1, 9): ("compute_pools", "accelerators_per_node"),
    (1, 4, 1, 10): ("compute_pools", "accelerator_type"),
    (1, 4, 1, 11): ("compute_pools", "cpu_type"),
    (1, 4, 1, 12): ("compute_pools", "clock_speed"),
    (1, 4, 1, 13): ("compute_pools", "clock_cycle_concurrency"),
    # Compute pool 2:
    (1, 4, 2, 1): ("compute_pools", "name"),
    (1, 4, 2, 2): ("compute_pools", "description"),
    ###(1, 4, 2, 3): ("compute_pools", "vendor"),  # TODO CIM V2.0 -> V2.2
    (1, 4, 2, 4): ("compute_pools", "model_number"),
    (1, 4, 2, 5): ("compute_pools", "number_of_nodes"),
    (1, 4, 2, 6): ("compute_pools", "memory_per_node"),
    (1, 4, 2, 7): ("compute_pools", "compute_cores_per_node"),
    # TODO CIM V2.0 -> V2.2, NIC component:
    ### (1, 4, 2, 8, 1, 1): ("compute_pools", "nic", "name"),
    ### (1, 4, 2, 8, 1, 2): ("compute_pools", "nic", "bandwidth"),
    ### (1, 4, 2, 8, 1, 3): ("compute_pools", "nic", "vendor"),
    ### (1, 4, 2, 8, 2, 1): ("compute_pools", "nic", "name"),
    ### (1, 4, 2, 8, 2, 2): ("compute_pools", "nic", "bandwidth"),
    ### (1, 4, 2, 8, 2, 3): ("compute_pools", "nic", "vendor"),
    ### (1, 4, 2, 8, 3, 1): ("compute_pools", "nic", "name"),
    ### (1, 4, 2, 8, 3, 2): ("compute_pools", "nic", "bandwidth"),
    ### (1, 4, 2, 8, 3, 3): ("compute_pools", "nic", "vendor"),
    (1, 4, 2, 9): ("compute_pools", "accelerators_per_node"),
    (1, 4, 2, 10): ("compute_pools", "accelerator_type"),
    (1, 4, 2, 11): ("compute_pools", "cpu_type"),
    (1, 4, 2, 12): ("compute_pools", "clock_speed"),
    (1, 4, 2, 13): ("compute_pools", "clock_cycle_concurrency"),
    # Storage pools...
    # Storage pool 1:
    (1, 5, 1, 1): ("storage_pools", "name"),
    # TODO CIM V2.0 -> V2.2:
    ### (1, 5, 1, 2): ("storage_pools", "file_system_sizes"),
    (1, 5, 1, 3): ("storage_pools", "description"),
    (1, 5, 1, 4): ("storage_pools", "type"),
    (1, 5, 1, 5): ("storage_pools", "vendor"),
    # Storage pool 2:
    (1, 5, 2, 1): ("storage_pools", "name"),
    # TODO CIM V2.0 -> V2.2:
    ### (1, 5, 2, 2): ("storage_pools", "file_system_sizes"),
    (1, 5, 2, 3): ("storage_pools", "description"),
    (1, 5, 2, 4): ("storage_pools", "type"),
    (1, 5, 2, 5): ("storage_pools", "vendor"),
    # Interconnect:
    # TODO CIM V2.0 -> V2.2 *and* third level:
    ### (1, 6, 1): ("compute_pools", "interconnect", "name"),
    ### (1, 6, 2): ("compute_pools", "interconnect", "topology"),
    ### (1, 6, 3): ("compute_pools", "interconnect", "description"),
    ### (1, 6, 4): ("compute_pools", "interconnect", "vendor"),
    # Benchmark performance:
    # TODO CIM V2.0 -> V2.2:
    ### (1, 7, 1): ("peak_performance",),
    ### (1, 7, 2): ("linpack_performance",),
}


def get_ws_questions_to_input_cells_mapping():
    """Return the mapping between question labels and worksheet input cells."""
    input_labels = WS_QUESTIONS_TO_INPUT_CELLS_MAPPING.copy()

    # Add WS questions corresponding to a second compute and storage pool
    second_pool_qs = {}
    for q_prefix in [COMPUTE_POOL_2_Q_NOS, STORAGE_POOL_2_Q_NOS]:
        previous_q_prefix = list(q_prefix[:-1]) + [q_prefix[-1] - 1]
        for q, val in input_labels.items():
            q_label_samesize = list(q[:len(q_prefix)])
            if q_label_samesize == previous_q_prefix:
                new_q = list(q)  # use original question label, but...
                new_q[:len(q_prefix)] = q_prefix  # ...change pool label 1 -> 2
                second_pool_qs[tuple(new_q)] = val  # and add this to the dict
    input_labels.update(second_pool_qs)  # add all new questions for 2nd pools

    # Extend the list with any potential labels for the questions
    applicable_models_q2 = {
        (1, 8, 2, N): 1 for N in range(
            1, MAX_NUMBER_MODELS_PER_INSTITUTE)
    }
    input_labels.update(applicable_models_q2)
    applicable_experiments_q2 = {
        (1, 9, 2, N): "SPECIAL CASE: 1+" for N in range(1, MAX_NUMBER_MIPS)}
    input_labels.update(applicable_experiments_q2)

    return input_labels


def get_machine_tabs(spreadsheet):
    """Return a list of all machine tab names in the machine worksheet."""
    all_machine_tabs = []

    # Don't rely on names having 'Machine X' format as that is not certain and
    # should not be relied upon: just exclude Frontis and example tab (first
    # two tabs) and assume rest are documented machines.
    for sheet in spreadsheet:
         if sheet.title not in ("Frontis", "Example"):
             all_machine_tabs.append(sheet)

    return all_machine_tabs


def find_input_cells(spreadsheet_tab, input_labels):
    """Find and return the input cells corresponding to the question labels."""
    label_values = {
        ".".join(str(subsec) for subsec in label): offset for label, offset in
        input_labels.items()
    }

    # Iterate over label first so stop check on rows via break once label found
    label_to_input_cell_mapping = {}
    for label, offset in label_values.items():
        for row in spreadsheet_tab.iter_rows(
                min_row=1, max_row=MAX_ROW, max_col=2):
            label_cell_value = str(row[LABEL_COLUMN].value).strip(" *")
            if label_cell_value == label:
                # Handle special cases of input cell(s) offsets:
                if isinstance(offset, str):
                    case = offset.lstrip("SPECIAL CASE: ")
                    logger.log_warning(
                        "Treating a special case for the offset of:",
                        label, "with rule:", case
                    )
                    if case.endswith("+"):
                        offsets = []
                        check_cell_at_offset = int(case.rstrip("+"))

                        # For N+, take all cells from N onwards until reach
                        # the first empty one, then stop:
                        while spreadsheet_tab.cell(
                                row[INPUT_COLUMN].row + check_cell_at_offset,
                                column=INPUT_COLUMN + 1
                        ).value:
                            offsets.append(check_cell_at_offset)
                            check_cell_at_offset += 1
                    else:  # not contiguous multiple input cells, other case
                        offsets = case.split("+")

                    # If offsets == [] here, no answer was provided so set as
                    # the input cell only the one default input box which will
                    # be recognised as empty later
                    if not offsets:
                        offsets = [check_cell_at_offset]

                    # Institutes may, against advice, have left some
                    # experiment input cells blank to indicate no applicable
                    # experiments by MIP, so to cater for these cases, replace
                    # empty values with 'NONE'
                    if label.startswith("1.9.2.") and not offsets:
                        offsets = [1]

                    # Now add the multiple offsets as a list, then move on:
                    label_to_input_cell_mapping[label] = [
                        row[INPUT_COLUMN].row + int(offset)
                        for offset in offsets
                    ]

                    break

                # Otherwise it is a simple offset, apply it:
                input_box_row = row[INPUT_COLUMN].row + offset
                label_to_input_cell_mapping[label] = input_box_row
                break

        # Remove inapplicable numbers for MIPs and experiments:
        if not label_to_input_cell_mapping.get(label, False):
            if label.startswith("1.9.2.") or label.startswith("1.8.2."):
                # Numbers not valid in this case, too few objects, so it
                # we can just skip these...
                logger.log_warning(
                    "Inapplicable model or MIP number skipped:", label)


    return label_to_input_cell_mapping


def extract_inputs_at_input_cells(input_cells, spreadsheet_tab):
    """Extract and return as a list values at the given input cells."""
    values = []
    if not isinstance(input_cells, list):
        input_cells = [input_cells]
    for index, input_cell in enumerate(input_cells):
        user_input = spreadsheet_tab.cell(
            row=input_cell, column=INPUT_COLUMN + 1).value

        # Distinguish from Falsy values e.g. False and 0 as user input
        if user_input is None:
            # Use a placeholder/default that is unlikely to be specified
            # (not None, False, etc.) to avoid potential clash with any
            # user input, for keeping track of input cells left empty.
            values.append(EMPTY_CELL_MARKER)
        else:
            values.append(str(user_input))
    return values


def get_top_cell_model_or_exp_name(input_cells, spreadsheet_tab):
    """Return the name of the topmost model or experiment input to the cells.

    In all such cases, the model or experiment name is at an offset of
    zero above the first input cell, i.e. directly above it, so note the
    logic here relies on this so must be adapated to generalise further.

    """
    if not isinstance(input_cells, list):
        input_cells = [input_cells]
    lowest_row_input_cell = min(input_cells)
    name = spreadsheet_tab.cell(
        row=lowest_row_input_cell - 1, column=INPUT_COLUMN + 1).value

    return name


def convert_tab_to_dict(spreadsheet_tab):
    """Return the full dictionary of inputs extracted from a machine tab."""
    all_input_cells = find_input_cells(
        spreadsheet_tab, get_ws_questions_to_input_cells_mapping())

    final_dict = {}
    for label, input_cell_or_cells in all_input_cells.items():
        if not isinstance(input_cell_or_cells, list):
            user_input = spreadsheet_tab.cell(
                row=input_cell_or_cells, column=INPUT_COLUMN + 1).value

            # Distinguish from Falsy values e.g. False and "None" as user input
            if user_input is None:
                # Use a placeholder/default that is unlikely to be specified
                # (not None, False, etc.) to avoid potential clash with any
                # user input, for keeping track of input cells left empty.
                final_dict[label] = EMPTY_CELL_MARKER
            else:  # else store the value, which may (rarely) be string "None"!
                # For cases where questions are populated based on institute
                # specific values (e.g. applicable models and exps questions),
                # also note these values for validation
                if label.startswith("1.8.2."):
                    name = get_top_cell_model_or_exp_name(
                        input_cell_or_cells, spreadsheet_tab)
                    final_dict[label] = {name: str(user_input)}
                else:
                    try:
                        final_dict[label] = str(user_input)
                    except:
                        # Python 2 only unicode-escape
                        logger.log_warning(
                            "WARNING: Python 2 only unicode issue with:",
                            label
                        )
                        final_dict[label] = user_input
        elif label.startswith("1.9.2."):
            if not spreadsheet_tab.cell(
                row=input_cell_or_cells[0], column=INPUT_COLUMN + 1).value:
                # Institutes may, against advice, have left some
                # experiment input cells blank to indicate no applicable
                # experiments by MIP, so to cater for these cases, replace
                # empty values with 'NONE'
                final_dict[label] = "NONE"
            else:
                name = get_top_cell_model_or_exp_name(
                    input_cell_or_cells, spreadsheet_tab)
                final_dict[label] = {name: extract_inputs_at_input_cells(
                    input_cell_or_cells, spreadsheet_tab)}
        else:
            final_dict[label] = extract_inputs_at_input_cells(
                input_cell_or_cells, spreadsheet_tab)

    return final_dict


def convert_str_type_to_cim_type(
        dicts_of_inputs, error_when_fail_type_conv=False, _print=False):
    """Convert a string to the type required by the CIM."""
    inputs_with_cim_type = []

    for input_dict in dicts_of_inputs:
        input_with_cim_type = {}
        # Filter out inputs where no value was specified, by marker
        submitted_inputs = {
            q_no: val for q_no, val in input_dict.items() if
            val != EMPTY_CELL_MARKER and val != [EMPTY_CELL_MARKER]
        }
        for q_no, q_answer in submitted_inputs.items():
            str_q_no = convert_question_number_str_to_tuple(q_no)
            # If the type is not correct it must be converted accordingly
            if str_q_no in WS_QUESTIONS_WITH_NON_STRING_TYPE:
                req_type = WS_QUESTIONS_WITH_NON_STRING_TYPE[str_q_no]
                if not isinstance(q_answer, req_type):
                    try:  # attempt conversion to correct CIM type
                        q_answer = req_type(q_answer)
                        if _print:
                            print("Converted {} from string to {}".format(
                                q_answer, req_type))
                    except (ValueError, TypeError):
                        if error_when_fail_type_conv:
                            raise TypeError(
                                "Input to question {} cannot be converted "
                                "to the required type {}: {}.".format(
                                    q_no, req_type, q_answer)
                            )
                        else:  # raise issues with group, so skip if not valid
                            if _print:
                                print(
                                    "WARNING: omitting answer to {} which "
                                    "could not be converted from string input "
                                    "to {} (not set to error): {}".format(
                                        q_no, req_type, q_answer)
                                )
                            continue
            input_with_cim_type[q_no] = q_answer
        inputs_with_cim_type.append(input_with_cim_type)

    return inputs_with_cim_type


def init_machine_cim(
        set_partition=False, two_compute_pools=True, two_storage_pools=True,
        online_docs_given=True
):
    """Initialise the CIM document for a CMIP6 Machine.

    Only up to two compute pools and storage pools may be specified.
    """
    kwargs = {
        "project": "CMIP6",
        "source": "spreadsheet",
        "version": 1,
        "institute": INSTITUTE
    }
    # Define the overall document which will be populated below
    machine_cim = pyesdoc.create(cim.Machine, **kwargs)

    # Connect the first-level properties to the top-level machine document
    if set_partition:
        machine_cim.partition = pyesdoc.create(cim.Partition, **kwargs)
    # TODO: list of given length, for now groups have all given len 1 answer
    if online_docs_given:
        machine_cim.online_documentation = [pyesdoc.create(
            cim.OnlineResource)]

    # Add pools based on the number required (max. two based on spreadsheet):
    if two_compute_pools:
        machine_cim.compute_pools = [
            pyesdoc.create(cim.ComputePool, **kwargs),
            pyesdoc.create(cim.ComputePool, **kwargs)
        ]
    else:
        machine_cim.compute_pools = [
            pyesdoc.create(cim.ComputePool, **kwargs)
        ]
    if two_storage_pools:
        machine_cim.storage_pools = [
            pyesdoc.create(cim.StoragePool, **kwargs),
            pyesdoc.create(cim.StoragePool, **kwargs)
        ]
    else:
        machine_cim.storage_pools = [
            pyesdoc.create(cim.StoragePool, **kwargs),
        ]

    return machine_cim


def convert_question_number_tuple_to_str(q_no):
    """Convert the tuple representing a question number into a string.

    Inverse to `convert_question_number_str_to_tuple`.
    """
    return ".".join([str(_int) for _int in q_no])


def convert_question_number_str_to_tuple(q_no):
    """Convert the string representing a question number into a tuple.

    Inverse to `convert_question_number_tuple_to_str`.
    """
    return tuple([int(_str) for _str in q_no.split(".")])


def get_inputs_and_mapping_to_cim(inputs_by_question_number_json):
    """Calculate the mapping of question numbers to CIM components."""
    questions_to_cim_mapping_str = {
        convert_question_number_tuple_to_str(q_no): val for q_no, val in
        QUESTIONS_TO_CIM_MAPPING.items()
    }
    return inputs_by_question_number_json, questions_to_cim_mapping_str


def set_cim_component(q_no, component, attribute_to_set, value_to_set):
    """Set components on the CIM document to register the question answer."""
    q_no_tuple = convert_question_number_str_to_tuple(q_no)
    if q_no_tuple in WS_QUESTIONS_WITH_ASSOCIATIONS:  # create an association
        cim_object = getattr(cim, WS_QUESTIONS_WITH_ASSOCIATIONS[q_no_tuple])
        # Set an association
        association = pyesdoc.associate_by_name(
            component, cim_object, value_to_set)
        setattr(component, attribute_to_set, association)
    else:
        # Set an attribute
        setattr(component, attribute_to_set, value_to_set)


def get_machine_doc(
        inputs_by_question_number_json, two_c_pools, two_s_pools, docs_given):
    """Create and return the completed CIM document for a CMIP6 Machine."""

    # Inititate machine CIM document
    # TODO: manage multiple partitions via set_partition flag kwarg
    machine_doc = init_machine_cim(
        two_compute_pools=two_c_pools, two_storage_pools=two_s_pools,
        online_docs_given=docs_given
    )
    inputs, q_to_cim_mapping = get_inputs_and_mapping_to_cim(
        inputs_by_question_number_json)

    # Match submitted questions to their corresponding machine CIM
    # components and set them accordingly on the document object
    for q_no, q_answer in inputs.items():
        if q_no in q_to_cim_mapping:
            corr_cim_comp = q_to_cim_mapping[q_no]
            level = len(corr_cim_comp)

            # a) Top level comps
            if level == 1:
                comp = corr_cim_comp[0]
                if comp == "online_documentation":  # special case 1
                    set_cim_component(
                        q_no,
                        getattr(machine_doc, comp)[0],
                        "name", "Online documentation describing a machine"
                    )
                    set_cim_component(
                        q_no,
                        getattr(machine_doc, comp)[0],
                        "linkage", q_answer[0]
                    )
                elif comp == "when_used":  # special case 2
                    set_cim_component(
                        q_no, machine_doc, comp, "When used")
                    if q_answer[0] != EMPTY_CELL_MARKER:
                        setattr(
                            getattr(machine_doc, comp),
                            "start_date", q_answer[0]
                        )
                    if q_answer[1] != EMPTY_CELL_MARKER:
                        setattr(
                            getattr(machine_doc, comp),
                            "end_date", q_answer[1]
                        )
                else:
                    set_cim_component(
                        q_no, machine_doc, comp, q_answer)
            elif level == 2:  # b) second-level comps e.g. storage pool
                level_1_comp, level_2_comp = corr_cim_comp

                # Special cases where need to set on one of two list values
                if level_1_comp in ("compute_pools", "storage_pools"):
                    # Determine if this is the first or second pool, as can
                    # indicated by the third value in the question number
                    # (1 -> first => Python index 0, etc.)
                    pool_index = int(str(q_no.split(".")[2])) - 1

                    if level_2_comp == "memory_per_node":
                        # Deal with special case:
                        set_cim_component(
                            q_no,
                            getattr(machine_doc, level_1_comp)[pool_index],
                            level_2_comp, "Memory per node"
                        )
                        setattr(
                            getattr(
                                getattr(machine_doc, level_1_comp)[pool_index],
                                level_2_comp
                            ), "volume", q_answer
                        )
                    else:
                        # Set value on the correct pool in the length-two list
                        set_cim_component(
                            q_no,
                            getattr(machine_doc, level_1_comp)[pool_index],
                            level_2_comp, q_answer
                        )
                else:
                    set_cim_component(
                        q_no,
                        getattr(machine_doc, level_1_comp),
                        level_2_comp, q_answer
                    )
            else:
                ValueError(
                    "Machine CIM should not be more than two levels deep."
                )

    return machine_doc


def generate_intermediate_dict_outputs(machines_spreadsheet):
    """Generate an intermediate dictionary for all machines per institute."""
    intermediate_dict_outputs = []
    tabs = get_machine_tabs(machines_spreadsheet)
    for machine_tab in tabs:
        intermediate_dict_outputs.append(convert_tab_to_dict(machine_tab))
    return intermediate_dict_outputs


def generate_outputs(
        machine_dict, two_c_pools, two_s_pools, docs_given, _print=False):
    """Generate and return all relevant outputs from a machine worksheet."""
    # Get the machine CIM document and applicable models and experiments
    cim_doc = get_machine_doc(
        machine_dict, two_c_pools, two_s_pools, docs_given)
    models = get_applicable_models(machine_dict)
    exps = get_applicable_experiments(machine_dict)

    if _print:
        print(cim_doc, models, exps)

    return cim_doc, models, exps


def filter_out_excess_pool(intermediate_dicts, q_no_start):
    """Filter any excess storage and/or compute pools from the dictionary."""
    filtered_dicts = []

    # Determine if a second pool has been described
    second_pool_described = [False] * len(intermediate_dicts)
    for index, int_dict in enumerate(intermediate_dicts):
        for q_no, q_answer in int_dict.items():
            if (not q_no.startswith(
                    convert_question_number_tuple_to_str(q_no_start))):
                continue
            if (q_answer != EMPTY_CELL_MARKER and
                q_answer != ([EMPTY_CELL_MARKER])):
                second_pool_described[index] = True

        # If the second pool has not been provided, filter out that question
        if not second_pool_described[index]:
            filtered_dicts.append(
                {
                    q_no: q_ans for q_no, q_ans in int_dict.items() if
                    not q_no.startswith(convert_question_number_tuple_to_str(
                        q_no_start))
                }
            )
        else:
            filtered_dicts.append(int_dict)

    return filtered_dicts, second_pool_described


def get_applicable_models(intermediate_dict):
    """Return all models applicable to the given institute as a list."""
    model_appl_answers = {
        q_no: q_ans for (q_no, q_ans) in intermediate_dict.items()
        if q_no.startswith("1.8")
    }
    all_applicable = model_appl_answers.pop("1.8.1")
    models_with_appl = model_appl_answers.values()

    if all_applicable == "ALL":
        # Take all listed models, so take all model keys regardless of value
        applicable_models = models_with_appl.keys()
    elif all_applicable == "SOME":
        # In this case must filter out ones specified as not being applicable
        applicable_models = [
            appl.keys()[0] for appl in models_with_appl if
            appl.values()[0] == "YES"
        ]
    else:
        raise ValueError(
            "Invalid input to enum. with 'ALL' or 'SOME' option only.")

    return applicable_models


def get_applicable_experiments(intermediate_dict):
    """Return all experiments applicable to the given institute as a list."""
    exp_appl_answers = {
        q_no: q_ans for (q_no, q_ans) in intermediate_dict.items()
        if q_no.startswith("1.9")
    }
    all_applicable = exp_appl_answers.pop("1.9.1")
    exp_with_appl = exp_appl_answers.values()

    if all_applicable == "ALL":
        mips_to_exps = vocabs.get_applicable_mips_with_experiments(INSTITUTE)
        # Flatten this mapping to MIPs out to set of all relevant experiments
        applicable_exps = set()
        for exps in mips_to_exps.values():
            applicable_exps.update(set(exp for exp in exps))
    elif all_applicable == "SOME":
        # In this case must filter out ones specified as not being applicable
        given_exps = [
            exp for exp in exp_with_appl if exp is not "NONE"]
        applicable_exps = {
            appl.keys()[0]: appl.values()[0] for appl in given_exps if
            appl.values()[0] != ["NONE"]
        }
        # TODO for rigour, add test to check for any weird inputs, e.g. None
        # in a list with named experiments...
    else:
        raise ValueError(
            "Invalid input to enum. with 'ALL' or 'SOME' option only.")

    return applicable_exps


def get_all_qs_to_inputs_mapping_for_institute():
    """Return JSON mapping question numbers to inputs for all machines.

    Note: this function is to facilitate the creation of the second-stage
    performance spreadsheets based on the inputs to the machine
    spreadsheet, rather than towards the creation of the machine CIM.
    """
    inputs = convert_ws_to_inputs(WS_IN_PATH)[0]

    # Tag each dictionary of inputs from a tab, corresponding to a given
    # documented machine, with the machine name, to aid processing
    final_qs_to_inputs_mapping = {}
    for machine_dict in inputs:
        machine_name = machine_dict["1.1.1"]  # compulsory q => guaranteed key
        final_qs_to_inputs_mapping[machine_name] = machine_dict

    return final_qs_to_inputs_mapping


def convert_ws_to_inputs(ws_location):
    """Return all processed inputs for a given machine worksheet."""
    # Locate and open template
    open_spreadsheet = load_workbook(filename=ws_location)

    # Extract inputs to spreadsheet as outputs ready to add to the CIM
    inputs_dicts = generate_intermediate_dict_outputs(open_spreadsheet)

    # Close template as now have extracted the outputs from it
    open_spreadsheet.close()

    # If only one of the two possible slots for storage (compute) pool has
    # been filled out, remove the second storage (compute) pool
    c_pool_filtered_dicts, two_c_pools = filter_out_excess_pool(
        inputs_dicts, COMPUTE_POOL_2_Q_NOS)
    filtered_inputs_dicts, two_s_pools = filter_out_excess_pool(
        c_pool_filtered_dicts, STORAGE_POOL_2_Q_NOS)

    has_docs = [
        inputs_d.get("1.2.3") != [EMPTY_CELL_MARKER]
        for inputs_d in filtered_inputs_dicts
    ]

    # Convert string outputs to their CIM type where non-string e.g. numeric,
    # doing this before processing the inputs in case there is a string that
    # cannot be converted, indicating a validation issue early on.
    type_converted_inputs_dicts = convert_str_type_to_cim_type(
        filtered_inputs_dicts)

    return type_converted_inputs_dicts, two_c_pools, two_s_pools, has_docs


# Main entry point.
if __name__ == '__main__':
    inputs, two_c_pools, two_s_pools, has_docs = convert_ws_to_inputs(
        WS_IN_PATH)

    # Iterate over all machine tabs to get all sets of outputs
    for index, input_dict in enumerate(inputs):
        # Return machine doc with applicable models and experiments:
        cim_out, apply_models_out, appl_exp_out = generate_outputs(
            input_dict, two_c_pools=two_c_pools[index],
            two_s_pools=two_s_pools[index], docs_given=has_docs[index]
        )

        # Validate the CIM document - there should not be any errors
        if pyesdoc.is_valid(cim_out):
            print(
                "Complete: machine CIM document generated and is valid.")
        else:
            print("Machine CIM document generated is not valid:")
            for err in pyesdoc.validate(cim_out):
                print(err)

        # Test serialisation of the machine doc...
        j = pyesdoc.encode(cim_out, pyesdoc.constants.ENCODING_JSON)
        assert json.loads(j)
        assert isinstance(
            pyesdoc.decode(j, pyesdoc.constants.ENCODING_JSON), cim.Machine)

        x = pyesdoc.encode(cim_out, pyesdoc.constants.ENCODING_XML)
        assert isinstance(
            pyesdoc.decode(x, pyesdoc.constants.ENCODING_XML), cim.Machine)

        # CIM document is valid and can be encoded correctly, so ready to
        # store it in the specified location as JSON:
        pyesdoc.write(cim_out, CIM_OUT_PATH, encoding=encoding)
        print("Machine CIM document successfully written to filesystem.")
