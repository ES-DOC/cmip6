"""
.. module:: init_performance_xls.py
   :license: GPL/CeCIL
   :platform: Unix, Windows
   :synopsis: Initialises CMIP6 per-machine per-model performance spreadsheets.

.. moduleauthor::
   Sadie Bartholomew <sadie.bartholomew@ncas.ac.uk>

"""

import argparse
import os
import shutil

from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

import pyessv

from lib.utils import io_mgr, logger, vocabs, constants

from generate_cim_via_json import (
    get_applicable_models,
    get_applicable_experiments,
    get_institute_json_mapping,
)


MACHINE_PLACEHOLDER = "<machine name>"
MODEL_PLACEHOLDER = "<model name>"
REALM_PLACEHOLDER = "<realm name>"

AGGREGATE_WS_NAME = "Aggregate for <model name> on <machine name>"
REALM_WS_NAME = "<realm name> of <model name> on <machine name>"


# Define command line argument parser.
_ARGS = argparse.ArgumentParser(
    "Initialises CMIP6 per-machine, per-model performance spreadsheets.")
_ARGS.add_argument(
    "--institution-id",
    help="An institution identifier",
    dest="institution_id",
    type=str,
    default="all"
    )
_ARGS.add_argument(
    "--xls-template",
    help="Path to XLS template",
    dest="xls_template",
    type=str
    )


def copy_cell(sheet, cell_to_copy_to, cell_to_copy_from):
    """Apply the value and background style of a cell to another named cell."""
    sheet[cell_to_copy_to] = sheet[cell_to_copy_from].value
    sheet[cell_to_copy_to]._style = copy(sheet[cell_to_copy_from]._style)


def set_institute_name_in_xls(institution, spreadsheet):
    """Write institute name into all relevant worksheets and their titles."""
    long_name = institution.data["name"].encode()
    short_name = institution.canonical_name.upper().encode()

    # Set name in front page worksheet
    frontis_sheet = spreadsheet["Frontis"]
    frontis_sheet["B4"] = "{} ({})".format(
        institution.data["name"].encode(),  # long name
        institution.canonical_name.upper().encode(),  # short name
    )


def set_machine_name_in_xls(machine_name, spreadsheet):
    """Write machine name into all relevant worksheets and their titles."""

    # Set name in front page worksheet
    frontis_sheet = spreadsheet["Frontis"]
    frontis_sheet["B5"] = machine_name

    # Set name in title of aggregate worksheet
    aggregate_ws = spreadsheet[AGGREGATE_WS_NAME]
    aggregate_title = AGGREGATE_WS_NAME.replace(
        MACHINE_PLACEHOLDER, machine_name)
    aggregate_ws.title = aggregate_title

    # Set name in title of realm worksheet (copied for each realm)
    realm_ws = spreadsheet[REALM_WS_NAME]
    realm_title = REALM_WS_NAME.replace(MACHINE_PLACEHOLDER, machine_name)
    realm_ws.title = realm_title

    # Set name within cells inside the aggregate and realm template worksheets
    for cell in ["B1", "B9", "B13"]:
        aggregate_name_answer = aggregate_ws[cell].value
        aggregate_ws[cell] = aggregate_name_answer.replace(
            MACHINE_PLACEHOLDER, machine_name)
        realm_name_answer = realm_ws[cell].value
        realm_ws[cell] = realm_name_answer.replace(
            MACHINE_PLACEHOLDER, machine_name)

    # Must store and return these in this case so we can further update them
    # to replace the model name placeholder with the model name.
    return (aggregate_title, realm_title)


def set_model_name_in_xls(
        model_name, spreadsheet, aggregate_ws_name, realm_ws_name):
    """Write model name into all relevant worksheets and their titles."""
    # Set name in front page worksheet
    frontis_sheet = spreadsheet["Frontis"]
    frontis_sheet["B6"] = model_name

    # Set name in title of aggregate worksheet
    aggregate_ws = spreadsheet[aggregate_ws_name]
    aggregate_title = aggregate_ws_name.replace(MODEL_PLACEHOLDER, model_name)
    aggregate_ws.title = aggregate_title

    # Set name in title of realm worksheet (copied for each realm)
    realm_ws = spreadsheet[realm_ws_name]
    realm_title = realm_ws_name.replace(MODEL_PLACEHOLDER, model_name)
    realm_ws.title = realm_title

    # Set name within cells inside the aggregate and realm template worksheets
    for cell in ["B1", "B9", "B17"]:
        aggregate_name_answer = aggregate_ws[cell].value
        aggregate_ws[cell] = aggregate_name_answer.replace(
            MODEL_PLACEHOLDER, model_name)
        realm_name_answer = realm_ws[cell].value
        realm_ws[cell] = realm_name_answer.replace(
            MODEL_PLACEHOLDER, model_name)

    # Must store and return these in this case so we can refer to the
    # worksheets later (this requires knowledge of their titles).
    return (aggregate_title, realm_title)


def set_realm_name_in_xls(realm_name, spreadsheet, realm_ws_name):
    """Write realm name into all relevant worksheets and their titles."""
    # Set name in title of realm worksheet
    realm_ws = spreadsheet[realm_ws_name]
    realm_title = realm_ws_name.replace(REALM_PLACEHOLDER, realm_name)
    realm_ws.title = realm_title.rstrip(" Copy")

    # Set name within cells inside the aggregate and realm template worksheets
    for cell in ["B1", "B9", "B23"]:
        realm_name_answer = realm_ws[cell].value
        realm_ws[cell] = realm_name_answer.replace(
            REALM_PLACEHOLDER, realm_name)

    # Return new title to use for copying worksheets later
    return realm_title


def create_tab_for_all_realms(all_realm_names, spreadsheet, realm_ws_name):
    """Create one fully-formatted realm worksheet for every possible realm."""
    # Get the realm worksheet template ready to duplicate for each realm
    realm_ws_template = spreadsheet[realm_ws_name]

    for realm_name in all_realm_names:
        # Create a copy of the template...
        new_realm_ws = spreadsheet.copy_worksheet(realm_ws_template)
        # ... and customise the copy to the specific realm.
        # Don't strip 'Copy' from the end of the WS name until later to
        # avoid overriding the original template and trouble from that...
        set_realm_name_in_xls(realm_name, spreadsheet, new_realm_ws.title)

    # Finally delete the original placeholder worksheet which is still there
    spreadsheet.remove(realm_ws_template)


def formatted_applicable_experiments(json_mapping_to_questions):
    """Pre-format applicable experiments to use in a drop-down list."""
    appl_exps_dict = get_applicable_experiments(json_mapping_to_questions)

    # Applicable experiments are stored in a dict by MIP key, but only
    # need the set of all applicable experiments, so flatten to set
    set_of_appl_exps = {
        exp for exps in appl_exps_dict.itervalues() for exp in exps}

    return set_of_appl_exps


def set_applicable_experiments_in_xls(
        applicable_experiments, spreadsheet, aggregate_ws_name):
    """Write drop-down list of applicable experiments into aggregate tabs."""
    appl_exps_as_comma_delim_string = ", ".join(applicable_experiments)
    # Note this requires string enclosed in quotes internally
    cell_formula = '"{}"'.format(appl_exps_as_comma_delim_string)

    aggregate_ws = spreadsheet[aggregate_ws_name]
    # Clear the placeholder value in the answer cell
    aggregate_ws["B27"].value = None
    # Add the drop-down with the available options to that cell
    appl_exps_dropdown = DataValidation(
        type="list", formula1=cell_formula, allow_blank=True)
    aggregate_ws.add_data_validation(appl_exps_dropdown)
    appl_exps_dropdown.add("B27")


def get_all_cmip6_realm_names():
    """Return a list containing the names of all CMIP6 realms as strings."""
    cmip6_realms = pyessv.WCRP.cmip6.get_source_realms()
    return [r.name for r in cmip6_realms.terms]


def customise_performance_template(
        spreadsheet, institution_name, machine_name, model_name):
    """Write out input details to customise the performance template."""
    # Customise the template appropriately to the given institute:
    #    1. Set the applicable institute, machine and model names
    set_institute_name_in_xls(institution_name, spreadsheet)
    aggregate_ws_title, realm_ws_title = set_machine_name_in_xls(
        machine_name, spreadsheet)

    aggregate_ws_title, realm_ws_title = set_model_name_in_xls(
        model_name, spreadsheet, aggregate_ws_title, realm_ws_title)

    #    2. Set a list of all applicable experiments as drop-down
    #       list for question 1.1.5 for the 'aggregate'
    #       performance tabs.
    set_applicable_experiments_in_xls(institution_name, spreadsheet)

    #    3. Duplicate tabs and tag for every possible realm
    all_realm_names = get_all_cmip6_realm_names()
    create_tab_for_all_realms(all_realm_names, spreadsheet, realm_ws_title)


def _main(args):
    """Main entry point.

    """
    # Defensive programming.
    if not os.path.exists(args.xls_template):
        raise ValueError("XLS template file does not exist")

    # Take generic template ready to process with institute-specific info.
    template_name = args.xls_template

    # Write out a customised template file for every institute
    for institution in vocabs.get_institutes(args.institution_id):
        institute_json_map = get_institute_json_mapping(institution)
        for machine, machine_json_map in institute_json_map.items():
            all_models_run_on_machine = get_applicable_models(machine_json_map)
            appl_exps = formatted_applicable_experiments(machine_json_map)

            for model in all_models_run_on_machine:
                # Open the template and customise it to the specific loop vars
                generic_template = load_workbook(filename=template_name)
                customise_performance_template(
                    generic_template, institution, machine, model, appl_exps)

                # Close template and save customised XLS to a new XLS file
                generic_template.close()
                final_spreadsheet_name = (
                    "{}_performance_of_{}_on_{}_{}.xlsx".format(
                        constants.CMIP6_MIP_ERA, model,
                        institution.canonical_name, machine
                    )
                )
                generic_template.save(final_spreadsheet_name)

                # Place the file into the appropriate directory, ultimately
                # writing one file per machine and applicable model combination
                dest = io_mgr.get_performance_spreadsheet(
                    institution, machine, model)
                logger.log(
                    "moving xls file for {}".format(institution.raw_name))
                shutil.copy(final_spreadsheet_name, dest)


# Main entry point.
if __name__ == '__main__':
    _main(_ARGS.parse_args())
